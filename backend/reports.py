import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfgen import canvas

from backend.db import get_productos, get_transacciones

# --- EXCEL REPORT GENERATION ---

def generate_excel_report(filepath):
    wb = Workbook()
    
    # 1. Sheet: Stock Actual
    ws_stock = wb.active
    ws_stock.title = "Stock Actual"
    ws_stock.views.sheetView[0].showGridLines = True
    
    # Fetch Data
    productos = get_productos()
    
    # Styling definitions
    font_title = Font(name="Segoe UI", size=16, bold=True, color="1E293B")
    font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="64748B")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Segoe UI", size=10)
    font_alert = Font(name="Segoe UI", size=10, bold=True, color="991B1B")
    font_warning = Font(name="Segoe UI", size=10, bold=True, color="9A3412")
    
    fill_header = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo accent
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_alert = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Soft red
    fill_warning = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid") # Soft orange
    
    thin_border_side = Side(border_style="thin", color="CBD5E1")
    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # Title Block
    ws_stock["A1"] = "REPORTE DE INVENTARIO - STOCK ACTUAL"
    ws_stock["A1"].font = font_title
    ws_stock["A2"] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws_stock["A2"].font = font_subtitle
    
    # Headers
    headers_stock = ["ID Producto", "Nombre del Producto", "Categoría", "Stock Actual", "Stock Mínimo", "Estado"]
    for col_idx, header in enumerate(headers_stock, start=1):
        cell = ws_stock.cell(row=4, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_cell
    
    # Data Rows
    row_idx = 5
    for p in productos:
        # Determine status and styles
        qty = p['cantidad']
        min_qty = p['stock_minimo']
        
        if qty == 0:
            status = "Sin Stock"
            row_fill = fill_alert
            status_font = font_alert
        elif qty <= min_qty:
            status = "Stock Bajo"
            row_fill = fill_warning
            status_font = font_warning
        else:
            status = "Suficiente"
            row_fill = fill_zebra if row_idx % 2 == 0 else None
            status_font = font_body
            
        row_data = [
            p['id'],
            p['nombre'],
            p['categoria'],
            qty,
            min_qty,
            status
        ]
        
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_stock.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = border_cell
            
            # Apply background highlights
            if row_fill and col_idx in [4, 5, 6]:
                cell.fill = row_fill
                if col_idx == 6:
                    cell.font = status_font
            elif row_idx % 2 == 0:
                cell.fill = fill_zebra
                
            # Alignments
            if col_idx in [1, 4, 5]:
                cell.alignment = align_right
            elif col_idx == 6:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
        row_idx += 1
        
    # Auto-adjust column widths
    for col in ws_stock.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Skip title row for length calculation
            if cell.row in [1, 2]:
                continue
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws_stock.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # 2. Sheet: Historial de Movimientos
    ws_history = wb.create_sheet(title="Historial de Movimientos")
    ws_history.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_history["A1"] = "HISTORIAL DE ENTRADAS Y SALIDAS (DEPÓSITO)"
    ws_history["A1"].font = font_title
    ws_history["A2"] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws_history["A2"].font = font_subtitle
    
    headers_history = ["ID Transacción", "ID Producto", "Producto", "Categoría", "Tipo", "Cantidad", "Área de Destino", "Fecha"]
    for col_idx, header in enumerate(headers_history, start=1):
        cell = ws_history.cell(row=4, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_cell
        
    transacciones = get_transacciones(limit=1000)
    
    fill_entrada = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid") # Soft green
    fill_salida = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid") # Soft red
    font_entrada = Font(name="Segoe UI", size=10, bold=True, color="047857")
    font_salida = Font(name="Segoe UI", size=10, bold=True, color="B91C1C")
    
    row_idx = 5
    for t in transacciones:
        tipo_str = t['tipo'].capitalize()
        row_fill = fill_entrada if t['tipo'] == 'entrada' else fill_salida
        tipo_font = font_entrada if t['tipo'] == 'entrada' else font_salida
        
        # Parse datetime
        fecha_dt = datetime.strptime(t['fecha'], "%Y-%m-%d %H:%M:%S")
        fecha_str = fecha_dt.strftime("%d/%m/%Y %H:%M:%S")
        
        row_data = [
            t['id'],
            t['producto_id'],
            t['producto_nombre'],
            t['producto_categoria'],
            tipo_str,
            t['cantidad'],
            t['area_destino'] if t['area_destino'] else "-",
            fecha_str
        ]
        
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_history.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = border_cell
            
            # Highlight Tipo
            if col_idx == 5:
                cell.fill = row_fill
                cell.font = tipo_font
                cell.alignment = align_center
            else:
                if row_idx % 2 == 0:
                    cell.fill = fill_zebra
                
            # Alignments
            if col_idx in [1, 2, 6]:
                cell.alignment = align_right
            elif col_idx in [5, 8]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
        row_idx += 1
        
    for col in ws_history.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2]:
                continue
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws_history.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    wb.save(filepath)

# --- PDF REPORT GENERATION ---

# Canvas for custom page numbering & footer
class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            super().showPage()
        super().save()

    def draw_page_number(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 9)
        self.setFillColor(colors.HexColor("#64748B"))
        
        # Draw header (on pages > 1)
        if self._pageNumber > 1:
            self.setStrokeColor(colors.HexColor("#E2E8F0"))
            self.setLineWidth(0.5)
            self.line(54, 750, 558, 750)
            self.drawString(54, 755, "Reporte de Depósito - Inventario & Movimientos")
            self.drawRightString(558, 755, datetime.now().strftime("%d/%m/%Y"))
            
        # Draw footer on all pages
        self.setStrokeColor(colors.HexColor("#E2E8F0"))
        self.setLineWidth(0.5)
        self.line(54, 50, 558, 50)
        
        footer_text = f"Página {self._pageNumber} de {page_count}"
        self.drawRightString(558, 38, footer_text)
        self.drawString(54, 38, "Sistema de Gestión de Stock de Depósito - Confidencial")
        self.restoreState()

def generate_pdf_report(filepath):
    # Setup document
    doc = SimpleDocTemplate(
        filepath,
        pagesize=letter,
        leftMargin=54, # 0.75 in
        rightMargin=54,
        topMargin=72,  # 1 in
        bottomMargin=72
    )
    
    # Styles
    styles = getSampleStyleSheet()
    
    # Custom styles
    style_title = ParagraphStyle(
        'DocTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=24,
        leading=28,
        textColor=colors.HexColor("#1E293B"),
        spaceAfter=6
    )
    
    style_subtitle = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor("#64748B"),
        spaceAfter=20
    )
    
    style_h2 = ParagraphStyle(
        'SectionHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=14,
        leading=18,
        textColor=colors.HexColor("#4F46E5"),
        spaceBefore=15,
        spaceAfter=10,
        keepWithNext=True
    )
    
    style_body = ParagraphStyle(
        'TableBody',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#334155")
    )
    
    style_body_bold = ParagraphStyle(
        'TableBodyBold',
        parent=style_body,
        fontName='Helvetica-Bold'
    )
    
    style_header_cell = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=12,
        textColor=colors.white
    )
    
    story = []
    
    # --- Title Page / Section ---
    story.append(Paragraph("Reporte General del Depósito", style_title))
    story.append(Paragraph(f"Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M:%S')} - Estado Actual de Inventario y Movimientos", style_subtitle))
    
    # --- Summary Section ---
    productos = get_productos()
    total_productos = len(productos)
    stock_total = sum(p['cantidad'] for p in productos)
    productos_bajo_stock = sum(1 for p in productos if p['cantidad'] <= p['stock_minimo'])
    
    summary_data = [
        [
            Paragraph("<b>Total Productos:</b>", style_body), Paragraph(str(total_productos), style_body_bold),
            Paragraph("<b>Stock Total:</b>", style_body), Paragraph(str(stock_total), style_body_bold),
            Paragraph("<b>Bajo Stock:</b>", style_body), Paragraph(f"<font color='#B91C1C'><b>{productos_bajo_stock}</b></font>", style_body_bold)
        ]
    ]
    summary_table = Table(summary_data, colWidths=[100, 50, 100, 50, 100, 104])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#F1F5F9")),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor("#E2E8F0")),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 15))
    
    # --- Stock Table ---
    story.append(Paragraph("Inventario de Stock Actual", style_h2))
    
    table_data = [[
        Paragraph("ID", style_header_cell),
        Paragraph("Producto", style_header_cell),
        Paragraph("Categoría", style_header_cell),
        Paragraph("Stock", style_header_cell),
        Paragraph("Mín.", style_header_cell),
        Paragraph("Estado", style_header_cell)
    ]]
    
    for p in productos:
        qty = p['cantidad']
        min_qty = p['stock_minimo']
        
        if qty == 0:
            status_html = "<font color='#991B1B'><b>Sin Stock</b></font>"
        elif qty <= min_qty:
            status_html = "<font color='#9A3412'><b>Bajo Stock</b></font>"
        else:
            status_html = "<font color='#047857'>Suficiente</font>"
            
        table_data.append([
            Paragraph(str(p['id']), style_body),
            Paragraph(p['nombre'], style_body),
            Paragraph(p['categoria'], style_body),
            Paragraph(str(qty), style_body),
            Paragraph(str(min_qty), style_body),
            Paragraph(status_html, style_body)
        ])
    
    # Width calculation (total printable width: 504 points)
    stock_table = Table(table_data, colWidths=[35, 179, 110, 50, 50, 80])
    
    # Dynamic styling for low stock rows
    table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4F46E5")), # Indigo Header
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
    ]
    
    # Row colors (zebra background)
    for idx in range(1, len(table_data)):
        p = productos[idx-1]
        qty = p['cantidad']
        min_qty = p['stock_minimo']
        if qty == 0:
            table_style.append(('BACKGROUND', (0, idx), (-1, idx), colors.HexColor("#FEE2E2"))) # Alert red
        elif qty <= min_qty:
            table_style.append(('BACKGROUND', (0, idx), (-1, idx), colors.HexColor("#FFEDD5"))) # Alert orange
        elif idx % 2 == 0:
            table_style.append(('BACKGROUND', (0, idx), (-1, idx), colors.HexColor("#F8FAFC"))) # Zebra stripe
            
    stock_table.setStyle(TableStyle(table_style))
    story.append(stock_table)
    story.append(Spacer(1, 20))
    
    # --- Transactions Table ---
    story.append(Paragraph("Últimos Movimientos de Depósito (Entradas y Salidas)", style_h2))
    
    trans_data = [[
        Paragraph("ID", style_header_cell),
        Paragraph("Producto", style_header_cell),
        Paragraph("Tipo", style_header_cell),
        Paragraph("Cant.", style_header_cell),
        Paragraph("Área Destino", style_header_cell),
        Paragraph("Fecha", style_header_cell)
    ]]
    
    transacciones = get_transacciones(limit=30) # Top 30 recent items for the PDF
    
    for t in transacciones:
        tipo_str = t['tipo'].capitalize()
        tipo_color = "#047857" if t['tipo'] == 'entrada' else "#B91C1C"
        tipo_html = f"<font color='{tipo_color}'><b>{tipo_str}</b></font>"
        
        fecha_dt = datetime.strptime(t['fecha'], "%Y-%m-%d %H:%M:%S")
        fecha_str = fecha_dt.strftime("%d/%m/%Y %H:%M")
        
        trans_data.append([
            Paragraph(str(t['id']), style_body),
            Paragraph(t['producto_nombre'], style_body),
            Paragraph(tipo_html, style_body),
            Paragraph(str(t['cantidad']), style_body),
            Paragraph(t['area_destino'] if t['area_destino'] else "-", style_body),
            Paragraph(fecha_str, style_body)
        ])
        
    trans_table = Table(trans_data, colWidths=[35, 179, 65, 45, 100, 80])
    
    trans_table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E293B")), # Slate Header
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
    ]
    
    for idx in range(1, len(trans_data)):
        if idx % 2 == 0:
            trans_table_style.append(('BACKGROUND', (0, idx), (-1, idx), colors.HexColor("#F8FAFC")))
            
    trans_table.setStyle(TableStyle(trans_table_style))
    story.append(trans_table)
    
    # Build Document using our custom NumberedCanvas
    doc.build(story, canvasmaker=NumberedCanvas)
